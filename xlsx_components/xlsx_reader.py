from openpyxl import load_workbook
from datetime import datetime, timedelta

from models import Student, Thesis, Employee, Slot


class XlsxReader:
    def __init__(self, file):
        self.workbook = load_workbook(filename=file)
        self.worksheet = self.workbook.active

        self.objects_to_read = {}

    def find_starting_row_and_column(self, title):
        for row in self.worksheet.iter_rows(max_row=10):
            for cell in row:
                if cell.value == title:
                    if self.worksheet.cell(row=cell.row + 1, column=cell.column).value is not None:
                        return cell.row + 1, cell.column
                    else:
                        row = cell.row
                        while True:
                            row += 1
                            if self.worksheet.cell(row=row + 1, column=cell.column).value is not None:
                                return row + 1, cell.column
                            else:
                                pass
        raise ValueError('Invalid file chosen.')

    def read_objects(self):
        for object_class, kwargs in self.objects_to_read.items():
            collection = []
            rows = []
            for attribute, title in kwargs.items():
                starting_row, kwargs[attribute] = self.find_starting_row_and_column(title)
                rows.append(starting_row)

            row = sorted(set(rows))[0]
            id_counter = 1

            while True:
                parameters = {}

                for attribute, column in kwargs.items():
                    parameters[attribute] = self.worksheet.cell(row=row, column=column).value

                if not all(value is None for value in parameters.values()):
                    parameters['id'] = id_counter
                    id_counter += 1

                    collection.append(
                        object_class(
                            parameters
                        )
                    )
                    row += 1
                else:
                    class_name = object_class.__name__
                    setattr(self, class_name.lower() if class_name.endswith('s') else class_name.lower() + 's',
                            collection)
                    break


class ThesisReader(XlsxReader):
    def __init__(self, file):
        super().__init__(file)

        self.objects_to_read = {
            Thesis: {
                'topic': 'Dyplom, temat',
                'individual': 'rodzaj rozlicz.',
                'faculty': 'Katedra - promotor',
                'supervisor': 'Dyplom, promotor',
                'reviewer': 'Dyplom, recenzent'
            },
            Student: {
                'name': 'Imi??',
                'surname': 'Nazwisko',
                'index': 'Nr albumu'
            }
        }

        self.thesis = []
        self.students = []

        self.read_objects()
        self.assign_thesis()

    def assign_thesis(self):
        for student in self.students:
            for i, thesis in enumerate(self.thesis):
                if thesis.topic.strip() == self.thesis[i - 1].topic.strip():
                    student.thesis = self.thesis[i - 1]
                    self.thesis.remove(thesis)
                else:
                    student.thesis = thesis

        for i, thesis in enumerate(self.thesis):
            thesis.id = i + 1

    def map_employees(self, employees):
        for thesis in self.thesis:
            supervisor = thesis.supervisor.split(' ')
            reviewer = thesis.reviewer.split(' ')
            for employee in employees:
                if f'{supervisor[-2]} {supervisor[-1]}' == f'{employee.name} {employee.surname}':
                    thesis.supervisor = employee
                elif f'{reviewer[-2]} {reviewer[-1]}' == f'{employee.name} {employee.surname}':
                    thesis.reviewer = employee


class EmployeesReader(XlsxReader):
    def __init__(self, file, slot_size=30, break_time=15, slot_block=5):
        super().__init__(file)
        self.objects_to_read = {
            Employee: {
                'name': 'imi?? i nazwisko',
                'surname': 'imi?? i nazwisko',
                'tenure': 'przewodnicz??cy-habilitacja',
                'online_slots': 'ONLINE slot czasowy na prac?? pojedyncz??;podw??jn??',
                'stationary_slots': 'STACJONARNIE slot czasowy na prac?? pojedyncz??;podw??jn??',
                'notes': 'uwagi'
            }
        }

        self.employees = []
        self.calendar = {}
        self.slots = {}

        self.read_objects()
        self.read_assign_availabilities()
        self.find_day_start_end()
        self.create_slots(slot_size=slot_size, break_=break_time, slot_block=slot_block)

        for employee in self.employees:
            self.assign_slots(employee)

    def assign_slots(self, employee):
        for day, slots in self.slots.items():
            for slot in slots:
                slot_start, slot_end = slot.start, slot.end
                availabilities = employee.availability[day].split(';')
                for availability in availabilities:
                    if availability.strip() == 'nie':
                        continue
                    avail_start, avail_end = (datetime.strptime(f'{day[0]} {x.strip()}', '%d %H:%M') for x in
                                              availability.split('-'))
                    if avail_start <= slot_start and avail_end >= slot_end:
                        employee.available_slots.append(slot)
                        break

    def read_assign_availabilities(self):
        availabilities = []
        starting_row, starting_column = self.find_starting_row_and_column('dost??pno????')
        days = [self.worksheet.cell(row=starting_row, column=starting_column).value]
        column = starting_column + 1
        while True:
            day = self.worksheet.cell(row=starting_row, column=column).value
            if day:
                days.append(day)
                column += 1
            else:
                break

        row = starting_row + 1
        while True:
            single_availability = {}
            for count, day in enumerate(days):
                single_availability[day] = self.worksheet.cell(row=row, column=starting_column + count).value

            if not all(x is None for x in single_availability.values()):
                availabilities.append(single_availability)
                row += 1
            else:
                break

        for employee, availability in zip(self.employees, availabilities):
            employee.availability = availability

    def find_day_start_end(self):
        for key in self.employees[0].availability.keys():
            hours = [
                hour for hours in [
                    employee.availability[key].split('-') for employee in self.employees if
                    employee.availability[key].strip() != 'nie'
                ] for hour in hours
            ]
            self.calendar[key] = '{}-{}'.format(min(hours), max(hours))

    def create_slots(self, slot_size, break_, slot_block):
        id_counter = 1
        for day, start_end_hours in self.calendar.items():
            start, end = (datetime.strptime(f'{day[0]} {x}', '%d %H:%M') for x in start_end_hours.split('-'))

            slots = []
            current_slot_in_block = 1
            current_end = start + timedelta(minutes=slot_size)
            while True:
                slots.append(Slot(day=day, start=start, end=current_end, id_=id_counter))
                start = current_end + timedelta(minutes=break_ if current_slot_in_block == slot_block else 0)
                current_end += timedelta(
                    minutes=slot_size + break_ if current_slot_in_block == slot_block else slot_size)
                id_counter += 1
                if current_end > end:
                    break
                elif current_slot_in_block != slot_block:
                    current_slot_in_block += 1
                else:
                    current_slot_in_block = 1

            self.slots[day] = slots
