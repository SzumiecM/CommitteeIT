import time

from openpyxl import load_workbook
from openpyxl.styles import Alignment


class XlsxWriter:
    def __init__(self, file):
        self.filename = file
        self.workbook = load_workbook(filename=self.filename)
        self.population = None
        self.algorithm_name = None

    def write(self, population, algorithm_name):
        self.population = population
        self.algorithm_name = algorithm_name
        self.copy_worksheet()
        self.write_individual()
        self.save()

    def copy_worksheet(self):
        work_sheet_name = f'{self.population.fitness} {self.algorithm_name}'
        self.worksheet = self.workbook.copy_worksheet(self.workbook.worksheets[0])
        self.worksheet.title = work_sheet_name

    def write_individual(self):
        head_of_committee_title = 'przewodniczący komisji egzaminu dyplomowego:'

        row, column = self.find_starting_row_and_column(head_of_committee_title)

        self.worksheet.cell(row - 1, column + 7).value = 'klucz sortowania'
        self.worksheet.cell(row - 1, column + 8).value = 'komentarz'

        for thesis in self.population.thesis:
            self.write_thesis(thesis, row, column)

            notes = []
            if thesis.head_of_committee.notes:
                notes.append(
                    f'{thesis.head_of_committee.name} {thesis.head_of_committee.surname}: {thesis.head_of_committee.notes}')
            for member in thesis.committee_members:
                if member.notes:
                    notes.append(f'{member.name} {member.surname}: {member.notes}')

            self.worksheet.cell(row, column + 8).alignment = Alignment(wrapText=True)
            self.worksheet.cell(row, column + 8).value = '\n'.join(notes) if notes else ''

            row += 1
            if not thesis.individual:
                self.write_thesis(thesis, row, column)
                row += 1

    def write_thesis(self, thesis, row, column):
        self.worksheet.cell(row, column).value = thesis.head_of_committee.__repr__()
        self.worksheet.cell(row, column + 1).value, self.worksheet.cell(row, column + 2).value = [x.__repr__() for x
                                                                                                  in
                                                                                                  thesis.committee_members]
        self.worksheet.cell(row, column + 3).value = thesis.slot.day
        self.worksheet.cell(row,
                            column + 4).value = f'{thesis.slot.start.hour}:{thesis.slot.end.minute if len(str(thesis.slot.end.minute))==2 else "00"} - {thesis.slot.end.hour}:{thesis.slot.end.minute if len(str(thesis.slot.end.minute))==2 else "00"}'
        self.worksheet.cell(row, column + 7).value = thesis.slot.id

    def save(self):
        while True:
            try:
                self.workbook.save(filename=self.filename)
                break
            except PermissionError:
                print('Could not save file')
                time.sleep(5)

    def find_starting_row_and_column(self, title):
        for row in self.worksheet.iter_rows(max_row=10):
            for cell in row:
                if cell.value == title:
                    if self.worksheet.cell(row=cell.row + 1, column=cell.column).value is None:
                        return cell.row + 1, cell.column
        raise ValueError('Invalid file chosen.')
