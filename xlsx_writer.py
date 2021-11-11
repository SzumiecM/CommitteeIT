from openpyxl import load_workbook


class XlsxWriter:
    def __init__(self, file):
        self.filename = file
        self.workbook = load_workbook(filename=self.filename)
        self.population = None

    def write(self, population):
        self.population = population
        self.copy_worksheet()
        self.write_thesis()
        self.save()

    def copy_worksheet(self):
        work_sheet_name = f'population {self.population.fitness}'
        self.worksheet = self.workbook.copy_worksheet(self.workbook.worksheets[0])
        self.worksheet.title = work_sheet_name

    def write_thesis(self):
        head_of_committee_title = 'przewodniczący komisji egzaminu dyplomowego:'
        # committee_member_title = 'członek komisji egzaminu dyplomowego:'
        # date_title = 'data'
        # hour_title = 'godzina'

        row, column = self.find_starting_row_and_column(head_of_committee_title)

        for thesis in self.population.thesis:
            self.worksheet.cell(row, column).value = thesis.head_of_committee.__repr__()
            self.worksheet.cell(row, column + 1).value, self.worksheet.cell(row, column + 2).value = [x.__repr__() for x
                                                                                                      in
                                                                                                      thesis.committee_members]
            self.worksheet.cell(row, column + 3).value = thesis.slot.day
            self.worksheet.cell(row,
                                column + 4).value = f'{thesis.slot.start.hour}:{thesis.slot.start.minute} - {thesis.slot.end.hour}:{thesis.slot.end.minute}'

            row += 1
            if not thesis.individual:
                self.worksheet.cell(row, column).value = thesis.head_of_committee.__repr__()
                self.worksheet.cell(row, column + 1).value, self.worksheet.cell(row, column + 2).value = [x.__repr__()
                                                                                                          for x in
                                                                                                          thesis.committee_members]
                self.worksheet.cell(row, column + 3).value = thesis.slot.day
                self.worksheet.cell(row,
                                    column + 4).value = f'{thesis.slot.start.hour}:{thesis.slot.start.minute} - {thesis.slot.end.hour}:{thesis.slot.end.minute}'
                row += 1

    def save(self):
        # todo replace later with same file name
        self.workbook.save(filename=self.filename)

    def find_starting_row_and_column(self, title):
        for row in self.worksheet.iter_rows(max_row=10):
            for cell in row:
                if cell.value == title:
                    if self.worksheet.cell(row=cell.row + 1, column=cell.column).value is None:
                        return cell.row + 1, cell.column
        raise Exception
